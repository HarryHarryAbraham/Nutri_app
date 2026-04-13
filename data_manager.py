import openpyxl
from openpyxl import load_workbook
from datetime import date, datetime
import os


NUTRITION_SHEET = "Nutrition_Data"
LOCAL_SHEET = "LocalData"

NUTRITION_COLS = {
    "A": "Health Center",
    "B": "Date (MM-DD-YYYY)",
    "C": "Patient Code",
    "D": "Beneficiary Category",
    "E": "Gender",
    "F": "Age of Lactating orPregnant (years)",
    "G": "Age of Child (months)",
    "H": "IDP/Host community",
    "I": "Z Score",
    "J": "MUAC (milimeter)",
    "K": "The final result",
    "L": "Action",
}


def get_col_index(letter):
    return ord(letter.upper()) - ord('A') + 1


def load_file(filepath):
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"الملف غير موجود: {filepath}")
    return load_workbook(filepath)


def save_file(wb, filepath):
    wb.save(filepath)


def search_patient(wb, code):
    """البحث عن مريض من الأسفل للأعلى — يعيد بيانات آخر زيارة"""
    ws = wb[NUTRITION_SHEET]
    code_str = str(code).strip()
    found_row_data = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] is not None and str(row[2]).strip() == code_str:
            found_row_data = row
    # آخر صف مطابق هو الأحدث
    return found_row_data


def get_last_row(wb):
    ws = wb[NUTRITION_SHEET]
    last = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row[:12]):
            last += 1
    return last


def save_visit(wb, data: dict):
    """حفظ زيارة جديدة في ورقة Nutrition_Data"""
    ws = wb[NUTRITION_SHEET]
    next_row = get_last_row(wb) + 1

    ws.cell(next_row, 1).value = data["center"]
    ws.cell(next_row, 2).value = data["date"]
    ws.cell(next_row, 3).value = data["code"]
    ws.cell(next_row, 4).value = data["category"]
    ws.cell(next_row, 5).value = data["gender"]
    ws.cell(next_row, 6).value = data["mother_age"]
    ws.cell(next_row, 7).value = data["child_age"]
    ws.cell(next_row, 8).value = data["idp"]
    ws.cell(next_row, 9).value = data["zscore"]
    ws.cell(next_row, 10).value = data["muac"]
    ws.cell(next_row, 11).value = data["diagnosis"]
    ws.cell(next_row, 12).value = data["action"]


def search_local(wb, patient_num):
    """البحث في LocalData بواسطة رقم المريض (عمود B)"""
    if LOCAL_SHEET not in wb.sheetnames:
        return None

    ws = wb[LOCAL_SHEET]
    search_val = str(patient_num).strip()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None and str(row[1]).strip() == search_val:
            return {
                "birthdate": row[0],
                "patient_num": row[1],
                "first_name": row[2] or "",
                "father_name": row[3] or "",
                "last_name": row[4] or "",
                "mother_name": row[5] or "",
                "clinic": row[6] or "",
                "consult_date": row[7],
            }
    return None


def export_nutrition_data(wb, dest_path):
    """تصدير ورقة Nutrition_Data إلى ملف Excel منفصل"""
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = NUTRITION_SHEET

    src_ws = wb[NUTRITION_SHEET]
    for row in src_ws.iter_rows(values_only=True):
        new_ws.append(list(row[:12]))

    new_wb.save(dest_path)


def calculate_age(birthdate):
    """حساب العمر بالتفصيل من تاريخ الميلاد"""
    if not birthdate:
        return ""
    if isinstance(birthdate, str):
        try:
            birthdate = datetime.strptime(birthdate, "%d/%m/%Y").date()
        except Exception:
            return ""
    if isinstance(birthdate, datetime):
        birthdate = birthdate.date()

    today = date.today()
    years = today.year - birthdate.year
    months = today.month - birthdate.month
    days = today.day - birthdate.day

    if days < 0:
        months -= 1
        from calendar import monthrange
        days += monthrange(today.year, today.month - 1 if today.month > 1 else 12)[1]
    if months < 0:
        years -= 1
        months += 12

    return f"{years} سنة و {months} شهر و {days} يوم"
